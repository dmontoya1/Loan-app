"""
Servicios de exportación de datos (PDF, Excel, CSV).
"""
from io import BytesIO
from decimal import Decimal
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class PDFExporter:
    """
    Exportador de datos a PDF.
    """
    
    @staticmethod
    def export_loans_pdf(loans, filename='prestamos.pdf'):
        """
        Exporta una lista de préstamos a PDF.
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        title = Paragraph(f"Reporte de Préstamos - {datetime.now().strftime('%d/%m/%Y')}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Tabla de datos
        data = [['Usuario', 'Monto', 'Interés', 'Cuotas', 'Pagadas', 'Estado']]
        
        for loan in loans:
            data.append([
                loan.user_profile.full_name,
                f"${loan.amount:,.2f}",
                f"{loan.interest_rate}%",
                str(loan.total_payments),
                f"{loan.completed_payments_count}/{loan.total_payments}",
                loan.get_status_display()
            ])
        
        table = Table(data, colWidths=[2.5*inch, 1*inch, 0.8*inch, 0.8*inch, 1*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Resumen
        total_loans = len(loans)
        total_amount = sum(loan.amount for loan in loans)
        active_loans = sum(1 for loan in loans if loan.status == 'ACTIVO')
        
        summary_data = [
            ['Total de Préstamos:', str(total_loans)],
            ['Préstamos Activos:', str(active_loans)],
            ['Monto Total Prestado:', f"${total_amount:,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(summary_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def export_payments_pdf(payments, filename='pagos.pdf'):
        """
        Exporta una lista de pagos a PDF.
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        title = Paragraph(f"Reporte de Pagos - {datetime.now().strftime('%d/%m/%Y')}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Tabla de datos
        data = [['Usuario', 'Préstamo', '# Pago', 'Monto', 'Fecha Venc.', 'Fecha Pago', 'Estado']]
        
        for payment in payments:
            data.append([
                payment.loan.user_profile.full_name,
                f"#{payment.loan.id}",
                str(payment.payment_number),
                f"${payment.amount:,.2f}",
                payment.due_date.strftime('%d/%m/%Y') if payment.due_date else '-',
                payment.payment_date.strftime('%d/%m/%Y') if payment.payment_date else '-',
                payment.get_status_display()
            ])
        
        table = Table(data, colWidths=[1.8*inch, 0.8*inch, 0.7*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (3, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Resumen
        total_payments = len(payments)
        total_amount = sum(payment.amount for payment in payments if payment.status == 'COMPLETADO')
        completed = sum(1 for payment in payments if payment.status == 'COMPLETADO')
        
        summary_data = [
            ['Total de Pagos:', str(total_payments)],
            ['Pagos Completados:', str(completed)],
            ['Monto Total Pagado:', f"${total_amount:,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(summary_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def export_payment_receipt_pdf(payment):
        """
        Genera un comprobante de pago en PDF estilizado.
        """
        buffer = BytesIO()
        # Márgenes un poco más amplios para el diseño
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=40)
        elements = []
        styles = getSampleStyleSheet()
        
        # Colores corporativos (Tailwind Cyan/Blue)
        primary_color = colors.HexColor('#22D3EE')
        dark_color = colors.HexColor('#1F2937')
        gray_color = colors.HexColor('#6B7280')
        light_gray = colors.HexColor('#F9FAFB')
        
        # Estilos
        title_style = ParagraphStyle(
            'ReceiptTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=dark_color,
            spaceAfter=5,
            alignment=TA_RIGHT,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'ReceiptSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=gray_color,
            spaceAfter=20,
            alignment=TA_RIGHT
        )
        
        company_name_style = ParagraphStyle(
            'CompanyName',
            parent=styles['Heading2'],
            fontSize=18,
            textColor=primary_color,
            spaceAfter=5,
            fontName='Helvetica-Bold'
        )
        
        section_header_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.white,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        # Cabecera (Empresa a la izquierda, "RECIBO" a la derecha)
        header_data = [
            [
                Paragraph(f"{payment.loan.tenant.name.upper() if hasattr(payment.loan, 'tenant') and payment.loan.tenant else 'SISTEMA DE PRÉSTAMOS'}", company_name_style),
                Paragraph("RECIBO DE PAGO", title_style)
            ],
            [
                Paragraph(f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']),
                Paragraph(f"Comprobante No. {payment.id:06d}", subtitle_style)
            ]
        ]
        
        header_table = Table(header_data, colWidths=[3.5*inch, 3.5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(header_table)
        
        # Línea separadora
        elements.append(Spacer(1, 0.2*inch))
        
        # Información del Cliente y Préstamo
        client_info_data = [
            [Paragraph("DATOS DEL CLIENTE", section_header_style), Paragraph("DATOS DEL PRÉSTAMO", section_header_style)],
            [
                f"Nombre: {payment.loan.user_profile.full_name}\n"
                f"DNI/Doc: {payment.loan.user_profile.document_number}\n"
                f"Teléfono: {payment.loan.user_profile.phone if hasattr(payment.loan.user_profile, 'phone') and payment.loan.user_profile.phone else 'N/A'}",
                f"Préstamo #: {payment.loan.id}\n"
                f"Fecha Inicio: {payment.loan.start_date.strftime('%d/%m/%Y')}\n"
                f"Monto Total Préstamo: ${payment.loan.amount:,.2f}"
            ]
        ]
        
        client_table = Table(client_info_data, colWidths=[3.4*inch, 3.4*inch])
        client_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), dark_color),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (1, 1), light_gray),
            ('BOX', (0, 0), (-1, -1), 1, colors.lightgrey),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        ]))
        
        elements.append(client_table)
        elements.append(Spacer(1, 0.4*inch))
        
        # Detalle del Pago
        elements.append(Paragraph("DETALLE DEL PAGO", ParagraphStyle('Detalle', parent=styles['Heading3'], textColor=dark_color, spaceAfter=10)))
        
        payment_date_str = payment.payment_date.strftime('%d/%m/%Y') if payment.payment_date else datetime.now().strftime('%d/%m/%Y')
        due_date_str = payment.due_date.strftime('%d/%m/%Y') if payment.due_date else '-'
        
        detail_data = [
            ['Descripción', 'Fecha Vencimiento', 'Fecha Pago', 'Total'],
            [f"Pago de Cuota {payment.payment_number} de {payment.loan.total_payments}", due_date_str, payment_date_str, f"${payment.amount:,.2f}"]
        ]
        
        detail_table = Table(detail_data, colWidths=[3.5*inch, 1.2*inch, 1.2*inch, 1*inch])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 15),
            ('LINEBELOW', (0, 0), (-1, 0), 2, dark_color),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.lightgrey),
        ]))
        
        elements.append(detail_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Totales
        totals_data = [
            ['Subtotal:', f"${payment.amount:,.2f}"],
            ['Total Pagado:', f"${payment.amount:,.2f}"]
        ]
        
        totals_table = Table(totals_data, colWidths=[5.4*inch, 1.5*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LINEABOVE', (0, 1), (1, 1), 1, dark_color),
        ]))
        
        elements.append(totals_table)
        elements.append(Spacer(1, 0.8*inch))
        
        # Firmas
        signatures_data = [
            ['_________________________', '_________________________'],
            ['Firma del Cliente', 'Firma Autorizada']
        ]
        
        signatures_table = Table(signatures_data, colWidths=[3.5*inch, 3.5*inch])
        signatures_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), gray_color),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        
        elements.append(signatures_table)
        
        # Nota pie de página
        elements.append(Spacer(1, 0.5*inch))
        note_style = ParagraphStyle(
            'ReceiptNote',
            parent=styles['Normal'],
            fontSize=8,
            textColor=gray_color,
            alignment=TA_CENTER
        )
        note = Paragraph("Este comprobante certifica que el pago ha sido recibido satisfactoriamente.<br/>Gracias por su puntualidad.", note_style)
        elements.append(note)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="comprobante_pago_{payment.id:06d}.pdf"'
        return response


class ExcelExporter:
    """
    Exportador de datos a Excel.
    """
    
    @staticmethod
    def export_loans_excel(loans, filename='prestamos.xlsx'):
        """
        Exporta una lista de préstamos a Excel.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Préstamos"
        
        # Estilos
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['ID', 'Usuario', 'Documento', 'Monto', 'Tasa Interés', 'Frecuencia', 
                   'Total Cuotas', 'Cuotas Pagadas', 'Estado', 'Fecha Inicio', 'Monto Total']
        ws.append(headers)
        
        # Estilo de encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for loan in loans:
            ws.append([
                loan.id,
                loan.user_profile.full_name,
                loan.user_profile.document_number,
                float(loan.amount),
                float(loan.interest_rate),
                loan.get_payment_frequency_display(),
                loan.total_payments,
                loan.completed_payments_count,
                loan.get_status_display(),
                loan.start_date.strftime('%d/%m/%Y') if loan.start_date else '',
                float(loan.total_amount)
            ])
        
        # Aplicar bordes y ajustar columnas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, (int, float)) and cell.column in [4, 5, 11]:
                    cell.number_format = '#,##0.00'
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 8, 'B': 25, 'C': 15, 'D': 12, 'E': 12, 'F': 12,
            'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Resumen en otra hoja
        ws_summary = wb.create_sheet("Resumen")
        ws_summary.append(['Resumen de Préstamos'])
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        total_loans = len(loans)
        total_amount = sum(float(loan.amount) for loan in loans)
        active_loans = sum(1 for loan in loans if loan.status == 'ACTIVO')
        
        summary_data = [
            ['Total de Préstamos:', total_loans],
            ['Préstamos Activos:', active_loans],
            ['Monto Total Prestado:', total_amount],
        ]
        
        ws_summary.append([])
        for row in summary_data:
            ws_summary.append(row)
            ws_summary[f'A{ws_summary.max_row}'].font = Font(bold=True)
            if isinstance(row[1], (int, float)):
                ws_summary[f'B{ws_summary.max_row}'].number_format = '#,##0.00'
        
        # Guardar
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def export_payments_excel(payments, filename='pagos.xlsx'):
        """
        Exporta una lista de pagos a Excel.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pagos"
        
        # Estilos
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['ID', 'Usuario', 'Préstamo #', '# Pago', 'Monto', 
                   'Fecha Vencimiento', 'Fecha Pago', 'Estado']
        ws.append(headers)
        
        # Estilo de encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for payment in payments:
            ws.append([
                payment.id,
                payment.loan.user_profile.full_name,
                payment.loan.id,
                payment.payment_number,
                float(payment.amount),
                payment.due_date.strftime('%d/%m/%Y') if payment.due_date else '',
                payment.payment_date.strftime('%d/%m/%Y') if payment.payment_date else '',
                payment.get_status_display()
            ])
        
        # Aplicar bordes y ajustar columnas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.column == 5:  # Columna de monto
                    cell.number_format = '#,##0.00'
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 8, 'B': 25, 'C': 10, 'D': 10, 'E': 12,
            'F': 15, 'G': 15, 'H': 12
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Resumen
        ws_summary = wb.create_sheet("Resumen")
        ws_summary.append(['Resumen de Pagos'])
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        total_payments = len(payments)
        completed = sum(1 for payment in payments if payment.status == 'COMPLETADO')
        total_amount = sum(float(payment.amount) for payment in payments if payment.status == 'COMPLETADO')
        
        summary_data = [
            ['Total de Pagos:', total_payments],
            ['Pagos Completados:', completed],
            ['Monto Total Pagado:', total_amount],
        ]
        
        ws_summary.append([])
        for row in summary_data:
            ws_summary.append(row)
            ws_summary[f'A{ws_summary.max_row}'].font = Font(bold=True)
            if isinstance(row[1], (int, float)):
                ws_summary[f'B{ws_summary.max_row}'].number_format = '#,##0.00'
        
        # Guardar
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def export_to_csv(data, filename='export.csv'):
        """
        Exporta datos a CSV (método simple).
        """
        import csv
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        if not data:
            return response
        
        writer = csv.writer(response)
        
        # Escribir encabezados (primera fila del primer elemento)
        if isinstance(data[0], dict):
            writer.writerow(data[0].keys())
            for row in data:
                writer.writerow(row.values())
        else:
            # Si es una lista de listas
            writer.writerows(data)
        
        return response
